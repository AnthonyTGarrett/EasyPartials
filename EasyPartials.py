#!/usr/bin/python3
# Author: Anthony Garrett
#
# Small script that will read in inventory location data from one spreadsheet
# and transfer that data to another spreadsheet
#
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

ORIGINAL_INPUT = "EXPORT.XLSX"
DATA_INPUT = "products.xlsx"


def main():

    # Loading the full pallet quantities from the product spreadsheet
    data_wb = load_workbook(DATA_INPUT)
    data_sheet = data_wb['Sheet1']

    data_products = data_sheet["A"]
    data_full_quantities = data_sheet["B"]
    data_products_dict = {}

    for i in range(len(data_products)):
        data_products_dict[data_products[i].value] = int(
            data_full_quantities[i].value)

    # Loading required data from totals spreadsheet

    wb = load_workbook(ORIGINAL_INPUT)
    partial_sheet = wb['Sheet1']

    # Deleting the first row which contains headers for the columns
    partial_sheet.delete_rows(1, 1)

    location_codes = partial_sheet["A"]
    products = partial_sheet["B"]
    actual_quantity = partial_sheet["D"]
    handling_unit = partial_sheet["G"]

    partials_list = []

    # Loop through actual_quantity list to compare actual quantity to total pallet quantity
    # of the specified product from data_products_dict
    for i in range(len(actual_quantity)):
        if actual_quantity[i].value < data_products_dict[int(products[i].value)]:
            partials_list.append(
                [location_codes[i].value, products[i].value, handling_unit[i].value, actual_quantity[i].value])

    partial_out_book = Workbook()

    partial_out_sheet = partial_out_book.active

    aisle_number = location_codes[0].value.split("-")[0]
    PARTIAL_OUTPUT_FILENAME = "Aisle-" + aisle_number + "-partials" + ".xlsx"

    # Setting Up headers for the spreadsheet
    partial_out_sheet["A1"] = "Storage Bin"
    partial_out_sheet.column_dimensions['A'].width = 12
    partial_out_sheet.row_dimensions[1].height = 25

    partial_out_sheet["B1"] = "Product"
    partial_out_sheet.column_dimensions['B'].width = 12

    partial_out_sheet["C1"] = "Handling Unit"
    partial_out_sheet.column_dimensions['C'].width = 20

    partial_out_sheet["D1"] = "Quantity"
    partial_out_sheet.column_dimensions['D'].width = 9

    for count, partial in enumerate(partials_list, start=2):
        partial_out_sheet["A" + str(count)] = partial[0]
        partial_out_sheet["B" + str(count)] = partial[1]
        partial_out_sheet["C" + str(count)] = partial[2]
        partial_out_sheet["D" + str(count)] = partial[3]

    for cell in partial_out_sheet['A:A']:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in partial_out_sheet['B:B']:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in partial_out_sheet['C:C']:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in partial_out_sheet['D:D']:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    partial_out_book.save(PARTIAL_OUTPUT_FILENAME)
    # os.remove(ORIGINAL_INPUT)


if __name__ == "__main__":
    main()
